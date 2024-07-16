import imaplib
import email
from email.header import decode_header
import openpyxl
import re
from langdetect import detect
from googletrans import Translator
import json
import os

# Load configuration from a file
def load_config(config_file):
    with open(config_file, 'r') as file:
        return json.load(file)

# Extract email address from the "From" header
def get_email_address(from_header):
    match = re.search(r'<(.+?)>', from_header)
    if match:
        return match.group(1)
    return from_header

# Function to check if the email body indicates the user does not need the license
def user_does_not_need_license(body):
    phrases = [
        "don't need the license",
        "do not need the license",
        "no longer need the license",
        "don't require the license",
        "do not require the license",
        "don't need adobe",
        "do not need adobe",
    ]
    for phrase in phrases:
        if phrase in body.lower():
            return True
    return False

def process_emails(config):
    # Connect to the email server
    mail = imaplib.IMAP4_SSL(config['imap_server'])
    mail.login(config['email'], config['password'])
    mail.select("inbox")  # Use the correct mailbox name

    # Search for all unread emails
    status, messages = mail.search(None, '(UNSEEN)')
    email_ids = messages[0].split()

    # Load the Excel workbook and select the active sheet
    wb = openpyxl.load_workbook(config['excel_file'])
    ws = wb.active

    # Initialize the Google Translate API
    translator = Translator()

    # Process each email
    for email_id in email_ids:
        status, msg_data = mail.fetch(email_id, "(RFC822)")
        raw_email = msg_data[0][1]
        msg = email.message_from_bytes(raw_email)
        
        # Decode email subject
        subject, encoding = decode_header(msg["Subject"])[0]
        if isinstance(subject, bytes):
            subject = subject.decode(encoding if encoding else "utf-8")
        
        # Get sender's email address
        from_header = msg["From"]
        sender_email = get_email_address(from_header)
        
        # Get email body
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get("Content-Disposition"))
                if "attachment" not in content_disposition and content_type == "text/plain":
                    body = part.get_payload(decode=True).decode()
                    break
        else:
            body = msg.get_payload(decode=True).decode()

        # Detect language and translate if necessary
        lang = detect(body)
        if lang != 'en':
            translated = translator.translate(body, src=lang, dest='en')
            body = translated.text
        
        # Check if the email is an OOO response, contains leave information, or states the user does not need the license
        if "out of office" in body.lower():
            status = "OOO"
            comment = "OOO auto-response"
            recommendation = ""
        elif "maternity leave" in body.lower() or "paternity leave" in body.lower():
            status = "On leave"
            comment = "On leave (maternity/paternity)"
            recommendation = ""
        elif user_does_not_need_license(body):
            status = "Response received"
            comment = "User does not need the license"
            recommendation = "Does not need license"
        else:
            status = "Response received"
            comment = "User responded"
            recommendation = "Needs license" if "need" in body.lower() else "Unclear, further review needed"
        
        # Update the Excel sheet (assuming email address in column A)
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value and sender_email in row[0].value:
                ws.cell(row=row[0].row, column=28, value=comment)  # Column AB
                ws.cell(row=row[0].row, column=29, value=status)   # Column AC
                ws.cell(row=row[0].row, column=30, value=recommendation)  # Column AD
                break

    # Save the updated Excel file
    wb.save(config['excel_file'])

    # Logout and close the connection
    mail.logout()

if __name__ == "__main__":
    config_file = os.getenv('CONFIG_FILE', 'config.json')
    config = load_config(config_file)
    process_emails(config)
