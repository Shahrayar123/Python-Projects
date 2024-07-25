import customtkinter as ctk
from tkinter import messagebox
from openpyxl import Workbook
import openpyxl
import pathlib
from datetime import datetime, timedelta

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

EMPLOYEES = ['FUNCIONARIO 1', 'FUNCIONARIO 2']
PASSWORD = '123456'

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.system()

    def layout_config(self):
        self.title('Cadastro de Ponto')
        self.geometry('300x300')

    def system(self):
        self.title_label = ctk.CTkLabel(self, text='Registro de Ponto', font=('Century Gothic Bold', 24), bg_color='transparent', text_color='#fff')
        self.title_label.place(x=50, y=50)

        self.option = ctk.CTkComboBox(self, values=['Entrada', 'Saída Intervalo', 'Volta Intervalo', 'Saída'], width=195, font=('Century Gothic', 16))
        self.option.set('Entrada')
        self.option.place(x=50, y=145)

        self.name = ctk.CTkComboBox(self, values=EMPLOYEES, width=195, font=('Century Gothic', 16))
        self.name.set(EMPLOYEES[0])
        self.name.place(x=50, y=100)

        self.punch_the_clock = ctk.CTkButton(self, text='Registrar Ponto', font=('Century Gothic', 16), width=195, command=self.submit)
        self.punch_the_clock.place(x=50, y=200)

        workbook_path = pathlib.Path('Registro de Ponto.xlsx')
        if not workbook_path.exists():
            workbook = Workbook()
            folha = workbook.active
            folha['A1'] = 'Nome Completo'
            folha['B1'] = 'Horário Entrada'
            folha['C1'] = 'Saída Intervalo'
            folha['D1'] = 'Volta Intervalo'
            folha['E1'] = 'Horário Saída'
            folha['F1'] = 'Dia'
            folha['G1'] = 'Horas Trabalhadas'
            folha['H1'] = 'Saldo de Horas'
            self.protect_sheet(folha)
            workbook.save('Registro de Ponto.xlsx')

    def protect_sheet(self, sheet):
        # Protege a planilha com uma senha
        sheet.protection.sheet = True
        sheet.protection.password = PASSWORD

    def submit(self):
        name = self.name.get()
        option = self.option.get()

        now = datetime.now()
        current_time = now.strftime('%H:%M:%S')
        current_day = now.strftime('%Y-%m-%d')

        workbook = openpyxl.load_workbook('Registro de Ponto.xlsx')
        folha = workbook.active

        if self.check_existing_entry(folha, name, current_day, option):
            messagebox.showerror("Erro", f"Já existe um registro para {option.lower()} no dia de hoje.")
            return

        found = False
        for row in range(2, folha.max_row + 1):
            if folha.cell(row=row, column=1).value == name and folha.cell(row=row, column=6).value == current_day:
                if option == 'Entrada' and not folha.cell(row=row, column=2).value:
                    folha.cell(row=row, column=2, value=current_time)
                elif option == 'Saída Intervalo' and not folha.cell(row=row, column=3).value:
                    folha.cell(row=row, column=3, value=current_time)
                elif option == 'Volta Intervalo' and not folha.cell(row=row, column=4).value:
                    folha.cell(row=row, column=4, value=current_time)
                elif option == 'Saída' and not folha.cell(row=row, column=5).value:
                    folha.cell(row=row, column=5, value=current_time)
                    self.calculate_worked_hours(folha, row)
                found = True
                break

        if not found:
            new_row = folha.max_row + 1
            folha.cell(column=1, row=new_row, value=name)
            if option == 'Entrada':
                folha.cell(column=2, row=new_row, value=current_time)
            elif option == 'Saída Intervalo':
                folha.cell(column=3, row=new_row, value=current_time)
            elif option == 'Volta Intervalo':
                folha.cell(column=4, row=new_row, value=current_time)
            elif option == 'Saída':
                folha.cell(column=5, row=new_row, value=current_time)
                self.calculate_worked_hours(folha, new_row)
            folha.cell(column=6, row=new_row, value=current_day)

        self.protect_sheet(folha)  # Protege a planilha novamente após a atualização
        workbook.save('Registro de Ponto.xlsx')
        messagebox.showinfo("Sistema", 'Dados salvos com sucesso!')
        self.clear()

    def check_existing_entry(self, folha, name, current_day, option):
        for row in range(2, folha.max_row + 1):
            if folha.cell(row=row, column=1).value == name and folha.cell(row=row, column=6).value == current_day:
                if option == 'Entrada' and folha.cell(row=row, column=2).value:
                    return True
                elif option == 'Saída Intervalo' and folha.cell(row=row, column=3).value:
                    return True
                elif option == 'Volta Intervalo' and folha.cell(row=row, column=4).value:
                    return True
                elif option == 'Saída' and folha.cell(row=row, column=5).value:
                    return True
        return False

    def calculate_worked_hours(self, folha, row):
        entrance = folha.cell(row=row, column=2).value
        lunch_time = folha.cell(row=row, column=3).value
        back_lunch_time = folha.cell(row=row, column=4).value
        leave = folha.cell(row=row, column=5).value

        if entrance and leave:
            entrance = datetime.strptime(entrance, '%H:%M:%S')
            leave = datetime.strptime(leave, '%H:%M:%S')

            if lunch_time and back_lunch_time:
                back_lunch_time = datetime.strptime(lunch_time, '%H:%M:%S')
                back_lunch_time = datetime.strptime(back_lunch_time, '%H:%M:%S')
                interval = back_lunch_time - lunch_time
            else:
                interval = timedelta()

            total_worked_hours = leave - entrance - interval
            folha.cell(row=row, column=7, value=str(total_worked_hours))

            worked_hours = total_worked_hours.total_seconds() / 3600
            balance = round(worked_hours - 7.20, 2)  
            folha.cell(row=row, column=8, value=f"{balance:.2f}")

    def clear(self):
        self.nome.set(EMPLOYEES[0])
        self.option.set('Entrada')

if __name__ == "__main__":
    app = App()
    app.mainloop()
